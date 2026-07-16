import base64
import io
import logging
import re

from odoo import _, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

# Fallback name used when a spreadsheet unit doesn't match anything already in the system -
# mirrors the Pakistani UoM naming used by data/pakistan_uom_data.xml (see that file for why
# "Sq.ft" on the sheet maps to the "ft²" uom.uom record, etc).
UOM_NAME_MAP = {
    "Sq.ft": "ft²",
    "Cft": "ft³",
    "Rft": "ft",
    "No.": "Units",
    "Bag": "Bag",
    "1,000 Bricks": "1,000 Bricks",
    "L.S.": "L.S. (Lump Sum)",
    "L.S": "L.S. (Lump Sum)",
    "Kg": "kg",
}

WORK_TYPE_CATEGORY_MAP = {
    "material": "material",
    "labour": "labour",
    "labor": "labour",
    "subcontract": "subcontract",
    "equipment": "equipment",
    "overhead": "overhead",
    "other": "other",
}

REQUIRED_HEADER_HINTS = {"S#", "Item / Material", "Stage / Work Head", "Unit"}


class ConstructionBoqImportWizard(models.TransientModel):
    _name = "construction.boq.import.wizard"
    _description = "Import BOQ Lines from an Excel Workbook"

    target = fields.Selection(
        [("template", "BOQ Template"), ("project_boq", "Project BOQ")],
        required=True, default="template",
        help="Import lines into a reusable BOQ Template, or directly into a project's BOQ "
             "(only while that BOQ is still in Draft - approved baselines are immutable).",
    )
    template_id = fields.Many2one(
        "construction.boq.template", string="Target Template",
        help="New lines are appended to this template's existing lines.",
    )
    boq_id = fields.Many2one(
        "construction.boq", string="Target Project BOQ",
        domain="[('state', '=', 'draft')]",
        help="Only Draft BOQs are shown - an approved/superseded baseline cannot be edited directly.",
    )
    file_data = fields.Binary(string="Excel File", required=True)
    file_name = fields.Char(string="File Name")

    # -- Helpers -------------------------------------------------------
    def _find_header_row(self, ws):
        for r in range(1, min(ws.max_row, 10) + 1):
            values = set()
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if val:
                    values.add(str(val).strip())
            if REQUIRED_HEADER_HINTS <= values:
                return r
        raise UserError(_(
            "Could not locate the BOQ header row (expected columns such as 'S#', "
            "'Item / Material', 'Stage / Work Head', 'Unit') within the first 10 rows of the "
            "first sheet."
        ))

    def _get_uom(self, raw_name, warnings):
        if not raw_name:
            return False
        raw_name = str(raw_name).strip()
        Uom = self.env["uom.uom"]
        mapped_name = UOM_NAME_MAP.get(raw_name)
        uom = False
        if mapped_name:
            uom = Uom.search([("name", "=", mapped_name)], limit=1)
        if not uom:
            uom = Uom.search([("name", "=", raw_name)], limit=1)
        if not uom:
            warnings.append(_("Unit '%s' was not found in the system - created a new generic unit for it.", raw_name))
            uom = Uom.create({
                "name": raw_name,
                "relative_uom_id": self.env.ref("uom.product_uom_unit").id,
                "relative_factor": 1.0,
            })
        return uom

    def _get_stage(self, raw_text, warnings):
        if not raw_text:
            return False
        text = str(raw_text).strip()
        text_clean = re.sub(r"^\d+[\s.\-]*", "", text).strip() or text
        Stage = self.env["construction.wbs.stage.template"]
        stage = Stage.search([("name", "=", text_clean)], limit=1)
        if not stage:
            stage = Stage.search([("name", "=", text)], limit=1)
        if not stage:
            stage = Stage.create({"name": text_clean})
            warnings.append(_("WBS stage '%s' was not found - created a new stage template for it.", text_clean))
        return stage

    def _get_work_type(self, raw_trade, warnings):
        if not raw_trade:
            return False
        trade = str(raw_trade).strip()
        category = WORK_TYPE_CATEGORY_MAP.get(trade.lower())
        WorkType = self.env["construction.work.type"]
        wt = False
        if category:
            wt = WorkType.search([("category", "=", category)], limit=1)
        if not wt:
            wt = WorkType.search([("name", "=ilike", trade)], limit=1)
        if not wt:
            warnings.append(_("Trade Type '%s' was not recognized - line imported without a Work Type.", trade))
        return wt

    # -- Import ----------------------------------------------------------
    def action_import(self):
        self.ensure_one()
        if self.target == "template" and not self.template_id:
            raise UserError(_("Select a target BOQ Template first."))
        if self.target == "project_boq":
            if not self.boq_id:
                raise UserError(_("Select a target Project BOQ first."))
            if self.boq_id.state != "draft":
                raise UserError(_("Lines can only be imported into a Project BOQ while it is in Draft state."))

        try:
            import openpyxl
        except ImportError:
            raise UserError(_("The 'openpyxl' Python library is required to import Excel files but is not installed."))

        workbook = openpyxl.load_workbook(io.BytesIO(base64.b64decode(self.file_data)), data_only=True)
        ws = workbook.worksheets[0]
        header_row = self._find_header_row(ws)

        headers = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=c).value
            if val:
                headers[str(val).strip()] = c

        def col(*names):
            for n in names:
                if n in headers:
                    return headers[n]
            return None

        c_item = col("Item / Material")
        c_stage = col("Stage / Work Head")
        c_unit = col("Unit")
        c_gross = col("Gross Qty")
        c_waste = col("Waste %")
        c_rate = col("B-Category Rate", "Rate")
        c_trade = col("Trade Type")
        c_basis = col("Technical Quantity Basis")
        c_remarks = col("Source / Remarks")

        if not c_item:
            raise UserError(_("Column 'Item / Material' was not found in the header row."))

        warnings = []
        imported = 0
        skipped = 0
        sequence = (max(self.template_id.line_ids.mapped("sequence") or [0])
                    if self.target == "template" else
                    max(self.boq_id.line_ids.mapped("sequence") or [0])) + 10

        Line = self.env["construction.boq.template.line"] if self.target == "template" else self.env["construction.boq.line"]

        wbs_by_stage_template = {}
        if self.target == "project_boq":
            for w in self.boq_id.project_id.wbs_ids:
                if w.stage_template_id:
                    wbs_by_stage_template[w.stage_template_id.id] = w

        for r in range(header_row + 1, ws.max_row + 1):
            item = ws.cell(row=r, column=c_item).value if c_item else None
            trade_raw = ws.cell(row=r, column=c_trade).value if c_trade else None
            is_subtotal = bool(trade_raw) and str(trade_raw).strip().lower() == "subtotal"

            if not item:
                # blank separator row - nothing to import
                skipped += 1
                continue

            vals = {"sequence": sequence}
            sequence += 10
            if self.target == "template":
                vals["template_id"] = self.template_id.id
            else:
                vals["boq_id"] = self.boq_id.id

            stage_raw = ws.cell(row=r, column=c_stage).value if c_stage else None
            stage = self._get_stage(stage_raw, warnings) if stage_raw else False

            if is_subtotal:
                vals["display_type"] = "subtotal"
                vals["name"] = str(item)
                vals["formula_type"] = "manual"
                if stage:
                    if self.target == "template":
                        vals["stage_template_id"] = stage.id
                    else:
                        wbs = wbs_by_stage_template.get(stage.id)
                        if wbs:
                            vals["wbs_id"] = wbs.id
                Line.create(vals)
                imported += 1
                continue

            vals["name"] = str(item)
            if c_basis:
                basis = ws.cell(row=r, column=c_basis).value
                if basis:
                    vals["technical_basis"] = str(basis)
            if c_remarks:
                remarks = ws.cell(row=r, column=c_remarks).value
                if remarks:
                    vals["source_remarks"] = str(remarks)

            unit_raw = ws.cell(row=r, column=c_unit).value if c_unit else None
            uom = self._get_uom(unit_raw, warnings)
            if uom:
                vals["uom_id"] = uom.id

            work_type = self._get_work_type(trade_raw, warnings) if trade_raw else False
            if work_type:
                vals["work_type_id"] = work_type.id
                if work_type.default_cost_code_id:
                    vals["cost_code_id"] = work_type.default_cost_code_id.id

            gross = ws.cell(row=r, column=c_gross).value if c_gross else 0.0
            vals["formula_type"] = "manual"
            vals["number"] = float(gross) if gross else 0.0

            waste = ws.cell(row=r, column=c_waste).value if c_waste else 0.0
            waste = float(waste) if waste else 0.0
            # Heuristic: sheets sometimes store waste as a fraction (0.03 = 3%) and sometimes
            # already as a percentage (3 = 3%). Values <= 1 are assumed to be fractions.
            if 0 < waste <= 1:
                waste *= 100.0
            vals["waste_pct"] = waste

            rate = ws.cell(row=r, column=c_rate).value if c_rate else 0.0
            vals["rate"] = float(rate) if rate else 0.0

            if stage:
                if self.target == "template":
                    vals["stage_template_id"] = stage.id
                else:
                    wbs = wbs_by_stage_template.get(stage.id)
                    if wbs:
                        vals["wbs_id"] = wbs.id
                    else:
                        warnings.append(_(
                            "No WBS phase found on this project for stage '%s' - line imported without a WBS link.",
                            stage.name,
                        ))

            Line.create(vals)
            imported += 1

        message = _("%(imported)s line(s) imported, %(skipped)s row(s) skipped.", imported=imported, skipped=skipped)
        if warnings:
            shown = warnings[:20]
            message += "\n" + "\n".join(shown)
            if len(warnings) > 20:
                message += "\n" + _("... and %s more warning(s) (see server log).", len(warnings) - 20)
            for w in warnings:
                _logger.warning("BOQ import: %s", w)

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("BOQ Import"),
                "message": message,
                "sticky": bool(warnings),
                "type": "warning" if warnings else "success",
            },
        }


class ConstructionBoqTemplateExport(models.Model):
    """Excel export counterpart to the import wizard above (FDD: BOQ templates must be
    round-trippable to/from the same spreadsheet layout used for import)."""

    _inherit = "construction.boq.template"

    def action_export_xlsx(self):
        self.ensure_one()
        try:
            import openpyxl
            from openpyxl.styles import Font
        except ImportError:
            raise UserError(_("The 'openpyxl' Python library is required to export Excel files but is not installed."))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = (self.name or "BOQ")[:31]

        headers = [
            "S#", "Category", "Stage / Work Head", "Item / Material", "Technical Quantity Basis",
            "Unit", "Gross Qty", "Waste %", "Net Qty", "Rate", "Amount", "Trade Type", "Source / Remarks",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        s_no = 0
        for line in self.line_ids.sorted("sequence"):
            if line.display_type:
                ws.append([
                    "", self.category_id.name or "", line.stage_template_id.name or "", line.name,
                    "", "", "", "", "", "", line.amount, "Subtotal", "",
                ])
                continue
            s_no += 1
            ws.append([
                s_no,
                self.category_id.name or "",
                line.stage_template_id.name or "",
                line.name,
                line.technical_basis or "",
                line.uom_id.name or "",
                line.gross_qty,
                line.waste_pct,
                line.net_qty,
                line.rate,
                line.amount,
                line.work_type_id.name or "",
                line.source_remarks or "",
            ])

        buf = io.BytesIO()
        wb.save(buf)
        attachment = self.env["ir.attachment"].create({
            "name": "%s.xlsx" % (self.name or "BOQ Template"),
            "type": "binary",
            "datas": base64.b64encode(buf.getvalue()),
            "res_model": self._name,
            "res_id": self.id,
        })
        return {
            "type": "ir.actions.act_url",
            "url": "/web/content/%s?download=true" % attachment.id,
            "target": "self",
        }
